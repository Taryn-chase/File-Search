#!/usr/bin/env python3
import os
import sys
import time
import glob
from pathlib import Path
import multiprocessing as mp
import PyPDF2
import openpyxl

# Function to prompt the user to choose a reference file from a given directory
def choose_reference_file(resources_dir, search_all=False):
    try:
        files = sorted(glob.glob(os.path.join(resources_dir, '*')))
        if not files:
            raise FileNotFoundError("No files found in the given directory.")
    except FileNotFoundError as e:
        print(f"Error: {e}")
        sys.exit(1)

    if search_all:
        return files
    else:
        while True:
            for idx, file in enumerate(files, 1):
                print(f"{idx}. {os.path.basename(file)}")
            try:
                choice = int(input("Please choose the Reference file: "))
                if 1 <= choice <= len(files):
                    return [files[choice - 1]]
                else:
                    print("Invalid selection. Please choose a valid file.")
            except ValueError:
                print("Invalid input. Please enter a number.")

# Function to search a given value in the reference file and append the matched lines to the search results file
def search_and_append(value, Reference_file, Search_results):
    print(f"Searching for value: {value} in {os.path.basename(Reference_file)}")
    
    file_extension = os.path.splitext(Reference_file)[1].lower()
    allowed_text_extensions = ['.txt', '.mnt', '.log']  # Add other extensions as needed

    if file_extension in allowed_text_extensions:
        try:
            with open(Reference_file, 'r', encoding='utf-8', errors='ignore') as ref_file, open(Search_results, 'a', encoding='utf-8', errors='ignore') as output:
                for line in ref_file:
                    if value in line:
                        output.write(line)
        except (FileNotFoundError, IOError) as e:
            print(f"Error: {e}")

    elif file_extension == '.pdf':
        try:
            with open(Search_results, 'a', encoding='utf-8', errors='ignore') as output:
                pdf_file = PyPDF2.PdfReader(Reference_file)
                for page_num in range(len(pdf_file.pages)):
                    page = pdf_file.pages[page_num]
                    text = page.extract_text()
                    if value in text:
                        output.write(f"Page {page_num + 1}:\n{text}\n")
        except (FileNotFoundError, IOError, PyPDF2._utils.PdfReadError) as e:
            print(f"Error: {e}")

    elif file_extension == '.xlsx':
        try:
            with open(Search_results, 'a', encoding='utf-8', errors='ignore') as output:
                wb = openpyxl.load_workbook(Reference_file, read_only=True, data_only=True)
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    for row in sheet.iter_rows():
                        row_values = [cell.value for cell in row]
                        if value in row_values:
                            output.write(f"Sheet: {sheet_name}, Row: {row_values}\n")
        except (FileNotFoundError, IOError, openpyxl.utils.exceptions.InvalidFileException) as e:
            print(f"Error: {e}")

    else:
        print(f"Unsupported file type: {file_extension}")

def main():
    # Define the file and directory paths
    values_file = os.path.expanduser('~/Documents/File_search/Search/Array.txt')
    resources_dir = os.path.expanduser('~/Documents/File_search/Resources')

    # Ask the user if they want to search against a single file or the entire directory
    while True:
        search_option = input("Choose an option:\n1. Search against a single file\n2. Search against the entire directory\n")
        if search_option == '1':
            search_all = False
            break
        elif search_option == '2':
            search_all = True
            break
        else:
            print("Invalid option. Please enter either 1 or 2.")

    try:
        Reference_files = choose_reference_file(resources_dir, search_all)
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)

    # Create a timestamp and output directory
    timestamp = time.strftime("%Y%m%d-%H%M%S")
    output_dir = os.path.expanduser(f'~/Documents/File_search/Output/output_{timestamp}')
    os.makedirs(output_dir, exist_ok=True)

    # Read values from the values_file
    try:
        with open(values_file, 'r') as f:
            values = [line.strip() for line in f]
    except (FileNotFoundError, IOError) as e:
        print(f"Error: {e}")
        sys.exit(1)

    # Present the user with options for searching
    print("\nChoose an option:")
    print("1. Search for values from the array (file)")
    print("2. Enter a single value to search for")
    option = input()

    # Record the start time of the search process
    start_time = time.time()

    # Perform search based on the user's choice
    if option == '1':
        for Reference_file in Reference_files:
            # Define the search results output file path
            output_file_name = f'{Path(Reference_file).stem}_{timestamp}_output.txt'
            Search_results = os.path.join(output_dir, output_file_name)

            with mp.Pool(processes=6) as pool:
                pool.starmap(search_and_append, [(value, Reference_file, Search_results) for value in values])
    elif option == '2':
        while True:
            single_value = input("Enter the value to search for (or type 'exit' to finish): ")
            if single_value == 'exit':
                break
            for Reference_file in Reference_files:
                # Define the search results output file path
                output_file_name = f'{Path(Reference_file).stem}_{timestamp}_output.txt'
                Search_results = os.path.join(output_dir, output_file_name)

                search_and_append(single_value, Reference_file, Search_results)
    else:
        print("Invalid option. Exiting.")
        sys.exit(1)

    # Show time elapsed
    end_time = time.time()
    time_elapsed = end_time - start_time
    print(f"Time elapsed: {time_elapsed} seconds")

    print("File(s) complete")
    # Show results in terminal
    with open(Search_results, 'r') as f:
        print(f.read())       

if __name__ == '__main__':
    main()
